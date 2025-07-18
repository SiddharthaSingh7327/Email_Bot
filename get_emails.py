import os
import sys
import time
import json
import pickle
import hashlib
import logging
import requests
import msal
import openpyxl
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from dataclasses import dataclass
from typing import Dict, List, Optional, Set
from datetime import datetime, timedelta

# ---------------- Logging Setup ----------------
def setup_logging():
    """Sets up logging to file and console."""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('lead_tracker.log'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ---------------- Configuration ----------------
# A NEW USER MUST EDIT THE VALUES IN THIS SECTION
@dataclass
class Config:
    CLIENT_ID: str = "91d3f9fe-f30d-4409-85fa-fa4a7c24c047"
    TENANT_ID: str = "64a9da10-e764-406f-a749-552dade47aa9"
    GEMINI_API_KEY: str = "AIzaSyCtecm-I_JzMVNtQHsAfzRykn1XbKwuPXU"
    
    # Do not change the lines below
    AUTHORITY: str = ""
    SCOPE: List[str] = None
    
    def __post_init__(self):
        self.AUTHORITY = f"https://login.microsoftonline.com/{self.TENANT_ID}"
        if self.SCOPE is None:
            self.SCOPE = [
                "User.Read",
                "Mail.Read",
                "Calendars.ReadWrite",
                "Files.ReadWrite.All" # Added for OneDrive folder creation
            ]

# ---------------- Data Models & File Management ----------------
@dataclass
class EmailData:
    id: str; sender_name: str; sender_email: str; subject: str; received: str; body_preview: str
    is_lead: bool = False
    parsed_data: Optional[Dict] = None
    calendar_event_id: Optional[str] = None

@dataclass
class OpportunityData:
    id: str; contact_name: str; company: str; email: str; phone: str
    opportunity_title: str; lead_status: str; notes: str; last_contacted: str; folder_link: str

class FileManager:
    """Handles loading and saving of pickle files for state management."""
    def load_pickle(self, filename: str) -> Set:
        if os.path.exists(filename):
            try:
                with open(filename, 'rb') as f: return pickle.load(f)
            except Exception as e: logger.error(f"Error loading {filename}: {e}")
        return set()
    
    def save_pickle(self, data: Set, filename: str):
        try:
            with open(filename, 'wb') as f: pickle.dump(data, f)
        except Exception as e: logger.error(f"Error saving {filename}: {e}")

# ---------------- Core Services ----------------
class AuthManager:
    """Handles Microsoft authentication and token management."""
    def __init__(self, config: Config):
        self.config = config
        self.cache = msal.SerializableTokenCache()
        self.token_cache_file = "token_cache.bin"
        if os.path.exists(self.token_cache_file):
            self.cache.deserialize(open(self.token_cache_file, "r").read())
        
        self.app = msal.PublicClientApplication(self.config.CLIENT_ID, authority=self.config.AUTHORITY, token_cache=self.cache)

    def get_access_token(self) -> Optional[str]:
        accounts = self.app.get_accounts()
        if accounts:
            result = self.app.acquire_token_silent(self.config.SCOPE, account=accounts[0])
            if result:
                self._save_cache(); return result.get("access_token")

        flow = self.app.initiate_device_flow(scopes=self.config.SCOPE)
        logger.info(f"🌐 {flow['message']}")
        result = self.app.acquire_token_by_device_flow(flow)

        if "access_token" in result:
            self._save_cache(); return result["access_token"]
        else:
            logger.error(f"Authentication failed: {result.get('error_description')}"); return None

    def _save_cache(self):
        if self.cache.has_state_changed:
            with open(self.token_cache_file, "w") as f: f.write(self.cache.serialize())

class GeminiParser:
    """Parses email content and generates summaries using the Gemini AI model."""
    def __init__(self, api_key: str):
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel("gemini-2.5-pro")
        
    def parse_email(self, email_body: str, email_subject: str) -> Optional[Dict]:
        prompt = f"""
        Analyze this email. First, decide if it is a sales lead. Ignore internal emails from @eucloid.com and automated replies.
        If it is NOT a lead, respond with: {{"is_lead": false}}
        If it IS a lead, analyze its content and respond in valid JSON with the following structure. Today is {datetime.now().strftime('%Y-%m-%d')}.
        {{"is_lead": true, "lead_status": "string (e.g., New Lead, Meeting Scheduled, Proposal Sent)", "has_meeting": boolean, "subject": "string", "date": "YYYY-MM-DD", "start_time": "HH:MM", "meeting_type": "string", "action_items": "string", "deadline": "YYYY-MM-DD"}}
        
        Email Subject: {email_subject}
        Email Body: {email_body[:1500]}
        """
        try:
            response = self.model.generate_content(prompt)
            clean_response = response.text.strip().replace("```json", "").replace("```", "")
            return json.loads(clean_response)
        except Exception as e:
            logger.error(f"Failed to parse email with Gemini: {e}"); return None

    def summarize_interactions(self, interaction_history: str) -> str:
        """Generates a summary of a lead's interaction history."""
        prompt = f"Based on the following email interaction log, provide a one-paragraph summary of the relationship with this lead so far.\n\n{interaction_history}"
        try:
            response = self.model.generate_content(prompt)
            return response.text.strip()
        except Exception as e:
            logger.error(f"Failed to generate summary with Gemini: {e}")
            return "Summary could not be generated."

class SharePointManager:
    """Handles interactions with OneDrive/SharePoint, like creating folders."""
    def __init__(self, headers: Dict):
        self.headers = headers
        self.base_url = "https://graph.microsoft.com/v1.0/me/drive/root"

    def create_folder_for_lead(self, opportunity_id: str) -> Optional[str]:
        """Creates a folder for a new lead and returns its web URL."""
        folder_name = f"Lead-{opportunity_id}"
        url = f"{self.base_url}/children"
        payload = {
            "name": folder_name,
            "folder": {},
            "@microsoft.graph.conflictBehavior": "rename"
        }
        try:
            response = requests.post(url, headers=self.headers, json=payload)
            response.raise_for_status()
            folder_data = response.json()
            logger.info(f"📁 Created OneDrive folder for Opportunity ID: {opportunity_id}")
            return folder_data.get("webUrl")
        except requests.exceptions.RequestException as e:
            logger.error(f"❌ Failed to create OneDrive folder for {opportunity_id}: {e}")
            return None

class CalendarManager:
    """Manages calendar event creation and duplicate detection."""
    def __init__(self, headers: Dict, file_manager: FileManager):
        self.headers = headers
        self.file_manager = file_manager
        self.processed_events = file_manager.load_pickle("processed_events.pkl")
    
    def create_event(self, parsed: Dict, sender_email: str, sender_name: str) -> Optional[str]:
        if not parsed or not parsed.get("has_meeting") or not parsed.get("date") or not parsed.get("start_time"):
            return None
        
        fingerprint = self._generate_event_fingerprint(parsed, sender_email)
        if fingerprint in self.processed_events:
            logger.info(f"Duplicate event detected, skipping: {parsed.get('subject')}"); return "duplicate_skipped"
        
        event_payload = self._build_event_payload(parsed, sender_email, sender_name)
        response = requests.post("https://graph.microsoft.com/v1.0/me/events", headers=self.headers, json=event_payload)
        
        if response.status_code == 201:
            event_id = response.json()["id"]
            self.processed_events.add(fingerprint)
            self.file_manager.save_pickle(self.processed_events, "processed_events.pkl")
            logger.info(f"Event created successfully: {parsed.get('subject')}"); return event_id
        else:
            logger.error(f"Failed to create event: {response.status_code} - {response.text}"); return None

    def _generate_event_fingerprint(self, parsed: Dict, email: str) -> str:
        data = f"{parsed.get('subject','')}|{parsed.get('date','')}|{parsed.get('start_time','')}|{email}".lower()
        return hashlib.md5(data.encode()).hexdigest()
    
    def _build_event_payload(self, parsed: Dict, email: str, name: str) -> Dict:
        start_dt_str = f"{parsed['date']}T{parsed['start_time']}:00"
        end_dt_obj = datetime.strptime(start_dt_str, "%Y-%m-%dT%H:%M:%S") + timedelta(hours=1)
        return {
            "subject": parsed.get("subject", "Meeting"),
            "start": {"dateTime": start_dt_str, "timeZone": "Asia/Kolkata"},
            "end": {"dateTime": end_dt_obj.strftime("%Y-%m-%dT%H:%M:%S"), "timeZone": "Asia/Kolkata"},
            "attendees": [{"emailAddress": {"address": email, "name": name}, "type": "required"}]
        }

# === LOCAL EXCEL REPORT GENERATOR (VISUALLY ENHANCED) ===
class ExcelReportGenerator:
    """Creates and updates a visually enhanced, local three-sheet Excel file."""
    def __init__(self, filename: str = "Opportunities.xlsx"):
        self.filename = Path(filename)

    def update_report(self, all_new_emails: List[EmailData], new_leads: List[EmailData], opportunities: Dict[str, OpportunityData]):
        """Creates or updates the local Excel file with three sheets and enhanced styling."""
        if not all_new_emails:
            logger.info("✅ No new emails to process.")
            return

        try:
            if self.filename.exists():
                wb = openpyxl.load_workbook(self.filename)
            else:
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
            
            self._ensure_sheets_exist(wb)
            
            all_emails_ws = wb["All Emails Log"]
            self._update_all_emails_log(all_emails_ws, all_new_emails)

            if new_leads:
                opps_ws = wb["Opportunities Master"]
                log_ws = wb["Interaction Log"]
                self._update_opportunities_sheet(opps_ws, opportunities)
                self._update_interaction_log_sheet(log_ws, new_leads)

            # Apply final styling to all sheets
            for ws in wb.worksheets:
                self._apply_styling(ws)

            wb.save(self.filename)
            logger.info(f"✅ Successfully updated {self.filename}.")
        except PermissionError:
            logger.error(f"❌ Error: Could not save to '{self.filename}'. Please close the file if it is open.")
        except Exception as e:
            logger.error(f"❌ An error occurred during Excel report generation: {e}", exc_info=True)

    def _ensure_sheets_exist(self, wb: Workbook):
        """Checks for all required sheets and creates them if they are missing."""
        if "Opportunities Master" not in wb.sheetnames:
            ws1 = wb.create_sheet("Opportunities Master", 0)
            headers = ["Opportunity ID", "Contact Name", "Company", "Email", "Phone", "Opportunity Title", "Lead Status", "Last Contacted", "Folder Link", "Notes"]
            ws1.append(headers)

        if "Interaction Log" not in wb.sheetnames:
            ws2 = wb.create_sheet("Interaction Log", 1)
            headers = ["Opportunity ID", "Meeting Date", "Meeting Summary", "Action Items", "Deadlines", "Meeting Type", "Timestamp"]
            ws2.append(headers)

        if "All Emails Log" not in wb.sheetnames:
            ws3 = wb.create_sheet("All Emails Log", 2)
            headers = ["Timestamp", "From", "Subject", "Is Lead?"]
            ws3.append(headers)

    def _apply_styling(self, ws):
        """Applies filters, frozen panes, column widths, and banding to a worksheet."""
        # Header styles based on sheet title
        header_fills = {
            "Opportunities Master": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            "Interaction Log": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "All Emails Log": PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
        }
        header_font = Font(bold=True, color="FFFFFF")
        
        # Alternating row color
        band_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Apply header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fills.get(ws.title, PatternFill())
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply banding and adjust column widths
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            if row_idx % 2 == 0: # Even rows
                for cell in row:
                    cell.fill = band_fill

        # Auto-fit columns
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50) # Cap width at 50
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Freeze header row and add filter
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    def _update_opportunities_sheet(self, ws, opportunities: Dict[str, OpportunityData]):
        """Updates the master sheet with new leads and the latest interaction data."""
        existing_opps = {ws.cell(row=i, column=1).value: i for i in range(2, ws.max_row + 1)}
        
        for opp_id, opp_data in opportunities.items():
            if opp_id in existing_opps:
                row_num = existing_opps[opp_id]
                ws.cell(row=row_num, column=7).value = opp_data.lead_status
                ws.cell(row=row_num, column=8).value = opp_data.last_contacted
                if opp_data.notes:
                    ws.cell(row=row_num, column=10).value = opp_data.notes
            else:
                new_row_data = [
                    opp_data.id, opp_data.contact_name, opp_data.company, opp_data.email,
                    opp_data.phone, opp_data.opportunity_title, opp_data.lead_status,
                    opp_data.last_contacted, f'=HYPERLINK("{opp_data.folder_link}", "Open Folder")', opp_data.notes
                ]
                ws.append(new_row_data)

    def _update_interaction_log_sheet(self, ws, new_leads: List[EmailData]):
        id_gen = lambda email: hashlib.md5(email.lower().encode()).hexdigest()[:8]
        for email in new_leads:
            if email.parsed_data and email.parsed_data.get("has_meeting"):
                parsed = email.parsed_data
                row_data = [
                    id_gen(email.sender_email),
                    parsed.get("date", ""),
                    email.body_preview[:1000],
                    parsed.get("action_items", ""),
                    parsed.get("deadline", ""),
                    parsed.get("meeting_type", "N/A"),
                    email.received
                ]
                ws.append(row_data)
    
    def _update_all_emails_log(self, ws, all_new_emails: List[EmailData]):
        for email in all_new_emails:
            row_data = [email.received, email.sender_email, email.subject, "Yes" if email.is_lead else "No"]
            ws.append(row_data)

    def get_interaction_history(self, opportunity_id: str) -> str:
        """Retrieves the full interaction history for a lead from the log."""
        if not self.filename.exists(): return ""
        wb = openpyxl.load_workbook(self.filename)
        if "Interaction Log" not in wb.sheetnames: return ""
        
        log_ws = wb["Interaction Log"]
        history = []
        for row in log_ws.iter_rows(min_row=2, values_only=True):
            if not opportunity_id or row[0] == opportunity_id:
                history.append(f"On {row[6]}, a meeting was logged: {row[2]}")
        return "\n".join(history)

# ---------------- Main Application ----------------
class LeadTracker:
    """Orchestrates the entire lead tracking workflow."""
    def __init__(self):
        self.config = Config()
        self.file_manager = FileManager()
        self.processed_emails = self.file_manager.load_pickle("processed_emails.pkl")
        self.auth_manager = AuthManager(self.config)
        self.gemini_parser = GeminiParser(self.config.GEMINI_API_KEY)
        self.report_generator = ExcelReportGenerator()
        self.access_token = None
        self.headers = None
        self.sharepoint_manager = None
        
    def initialize(self) -> bool:
        """Initializes services and gets an access token."""
        logger.info("🚀 Initializing Lead Tracker...")
        self.access_token = self.auth_manager.get_access_token()
        if not self.access_token:
            logger.error("❌ Failed to get access token"); return False
        self.headers = {"Authorization": f"Bearer {self.access_token}"}
        self.sharepoint_manager = SharePointManager(self.headers)
        logger.info(f"✅ Initialization successful. Loaded {len(self.processed_emails)} previously processed emails.")
        return True
    
    def process_emails(self, calendar_manager: CalendarManager) -> List[EmailData]:
        """Fetches, filters, and processes new emails."""
        all_newly_processed = []
        try:
            graph_url = "https://graph.microsoft.com/v1.0/me/messages?$top=25&$orderby=receivedDateTime desc"
            response = requests.get(graph_url, headers=self.headers, timeout=30)
            response.raise_for_status()
            messages = response.json().get("value", [])
            logger.info(f"📨 Emails fetched: {len(messages)}")

            for msg in messages:
                email_id = msg["id"]
                if email_id in self.processed_emails:
                    continue

                logger.info(f"\n📨 Checking email: '{msg.get('subject', '')}' from {msg.get('from', {}).get('emailAddress', {}).get('address', '')}")
                
                email_data = EmailData(
                    id=email_id,
                    sender_name=msg.get("from", {}).get("emailAddress", {}).get("name", ""),
                    sender_email=msg.get("from", {}).get("emailAddress", {}).get("address", ""),
                    subject=msg.get("subject", ""),
                    received=msg.get("receivedDateTime", ""),
                    body_preview=msg.get("body", {}).get("content", "")
                )
                
                parsed = self.gemini_parser.parse_email(email_data.body_preview, email_data.subject)
                
                if parsed and parsed.get("is_lead"):
                    logger.info("✅ ACCEPTED by Gemini as a lead.")
                    email_data.is_lead = True
                    email_data.parsed_data = parsed
                    if parsed.get("has_meeting"):
                        email_data.calendar_event_id = calendar_manager.create_event(parsed, email_data.sender_email, email_data.sender_name)
                else:
                    logger.info("❌ REJECTED by Gemini as not a lead.")

                all_newly_processed.append(email_data)
                self.processed_emails.add(email_id)

        except requests.exceptions.RequestException as e:
            logger.error(f"❌ Network error fetching emails: {e}")
        except Exception as e:
            logger.error(f"❌ Error processing emails: {e}")
            
        if all_newly_processed:
            self.file_manager.save_pickle(self.processed_emails, "processed_emails.pkl")
            
        return all_newly_processed

    def generate_opportunities(self, emails: List[EmailData]) -> Dict[str, OpportunityData]:
        """Generates a dictionary of unique opportunities from a list of emails."""
        opportunities = {}
        for email in emails:
            opp_id = hashlib.md5(email.sender_email.lower().encode()).hexdigest()[:8]
            
            # Get the most recent email for this opportunity in the current batch
            if opp_id not in opportunities or email.received > opportunities.get(opp_id, {}).get('last_contacted', ''):
                
                # Check if this is a brand new opportunity
                is_new = opp_id not in opportunities
                folder_link = self.sharepoint_manager.create_folder_for_lead(opp_id) if is_new else opportunities.get(opp_id, {}).get('folder_link', '')
                
                # Check if we need to generate an AI summary
                notes = ""
                history = self.report_generator.get_interaction_history(opp_id)
                if len(history.split('\n')) > 3: # Trigger summary after 3 interactions
                    notes = self.gemini_parser.summarize_interactions(history)

                opportunities[opp_id] = OpportunityData(
                    id=opp_id, contact_name=email.sender_name, company=email.sender_email.split('@')[1].split('.')[0].title(),
                    email=email.sender_email, phone="", opportunity_title=f"Opportunity with {email.sender_name}",
                    lead_status=email.parsed_data.get("lead_status", "New Lead") if email.parsed_data else "New Lead",
                    notes=notes,
                    last_contacted=email.received,
                    folder_link=folder_link
                )
        return opportunities

# ---------------- Main Execution ----------------
def run_continuously(interval_minutes=5):
    """Main execution loop that runs indefinitely."""
    app = LeadTracker()
    if not app.initialize():
        return
    
    calendar_manager = CalendarManager(app.headers, app.file_manager)
    report_generator = ExcelReportGenerator()
    last_summary_day = -1

    while True:
        today = datetime.now().weekday() # Monday is 0 and Sunday is 6
        
        # Weekly Summary Logic (runs on Fridays)
        if today == 4 and last_summary_day != today:
            history = report_generator.get_interaction_history("") # Get all for the week
            weekly_summary = f"Weekly Lead Summary:\n{len(history.splitlines())} new interactions logged this week."
            logger.info("="*60); logger.info(weekly_summary); logger.info("="*60)
            last_summary_day = today
        elif today != 4:
            last_summary_day = -1 # Reset day tracker

        logger.info(f"🔄 Starting new cycle at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        all_new_emails = app.process_emails(calendar_manager)
        
        if all_new_emails:
            new_leads = [email for email in all_new_emails if email.is_lead]
            logger.info(f"Found {len(new_leads)} new lead(s) out of {len(all_new_emails)} new email(s). Updating report...")
            
            if new_leads:
                opportunities = app.generate_opportunities(new_leads)
                report_generator.update_report(all_new_emails, new_leads, opportunities)
        else:
            logger.info("✅ No new emails to process.")
        
        logger.info(f"😴 Sleeping for {interval_minutes} minutes...")
        time.sleep(interval_minutes * 60)

if __name__ == "__main__":
    try:
        run_continuously(interval_minutes=5)
    except KeyboardInterrupt:
        logger.info("⚠️ Application interrupted by user. Shutting down.")
    except Exception as e:
        logger.error(f"❌ Critical application error: {e}", exc_info=True)
